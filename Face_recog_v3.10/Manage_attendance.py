from  openpyxl import Workbook , load_workbook , utils
import datetime
import calendar
import os 


def Add_name(name):
    filename = 'Attendance.xlsx'
    wb = load_workbook(filename=filename)
    month_yr = str(datetime.datetime.now()).split(" ")[0][:-3]
    # print(month_yr)
    ws = wb[month_yr]
    row_to_write = ws.max_row+1
    ws.cell(row=row_to_write, column=1).value = f"{name}"
    wb.save(filename=filename)
    print(row_to_write)

def mark_attendance():
  print('MArked')
