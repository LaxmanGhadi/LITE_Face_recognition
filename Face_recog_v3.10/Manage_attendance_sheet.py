from  openpyxl import Workbook , load_workbook , utils
import datetime
import calendar
import os 
# Check if excel sheet eand make if not found
wb = Workbook()
if  not os.path.exists("Attendance.xlsx"):
    ws = wb.active
    wb.save("Attendance.xlsx")
else:
    wb = load_workbook('Attendance.xlsx')
    latest_month = wb.sheetnames[-1]
    current_date = str(datetime.datetime.now()).split(' ')[0]# -> 2025-12-12 (YYYY-MM-DD)
    current_month_yr  = current_date[:-3] # -> 2025-12 (YYYY-MM)
    if not latest_month == current_month_yr:
        ns = wb.create_sheet(current_month_yr)
        year,month, day = map(int,current_date.split('-'))
        mo_days = calendar.monthrange(year=year, month=month)# No of days in a month
        for day in range(1,mo_days[1]+1):
            ns.cell(row = 1 , column = day+1).value = datetime.date(year=year, month=month, day=day)
            ns.column_dimensions[utils.get_column_letter(day+1)].width = 12
        
        people= os.listdir('Dataset/People')
        ns["A1"] = "Names"
        for idx,person in enumerate(people):
            ns.cell(row =  idx+2, column = 1).value = f"{person}"
        wb.save("Attendance.xlsx")
        print(mo_days[1])

    # else :
