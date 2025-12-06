from  openpyxl import Workbook , load_workbook , utils
import datetime
import calendar
import os 

class AttendanceManager:
  def __init__(self, filename='Attendance.xlsx'):
    self.filename = filename

    # Create workbook if not exists
    if not os.path.exists(self.filename):
        wb = Workbook()
        month_yr = str(datetime.datetime.now()).split(" ")[0][:-3]
        wb.create_sheet(title=month_yr)
        wb.save(self.filename)

  def make_month(current_month_yr,current_date):
    wb = load_workbook('Attendance.xlsx')
    ns = wb.create_sheet(current_month_yr)
    year,month, day = map(int,current_date.split('-'))
    mo_days = calendar.monthrange(year=year, month=month)# No of days in a month
    for day in range(1,mo_days[1]+1):
        ns.cell(row = 1 , column = day+1).value = datetime.date(year=year, month=month, day=day)
        ns.column_dimensions[utils.get_column_letter(day+1)].width = 12
    
    people= os.listdir('Dataset/People')
    ns["A1"] = "Names"
    for idx,person in enumerate(people):
        ns.cell(row =  idx+2, column = 1).value = f"{person}"
    wb.save("Attendance.xlsx")




  def get_latest_month(self):# Check for the latest worksheet and if the latest is not current month make a new sheet
    wb = load_workbook('Attendance.xlsx')
    latest_month = wb.sheetnames[-1]
    current_date = str(datetime.datetime.now()).split(' ')[0]# -> 2025-12-12 (YYYY-MM-DD)
    current_month_yr  = current_date[:-3] # -> 2025-12 (YYYY-MM)
    if not latest_month == current_month_yr:
      self.make_month(current_month_yr,current_date)
    return wb, wb[current_month_yr]

  def Add_name(self,name):
      filename = 'Attendance.xlsx'
      wb = load_workbook(filename=filename)
      month_yr = str(datetime.datetime.now()).split(" ")[0][:-3]
      # print(month_yr)
      ws = wb[month_yr]
      row_to_write = ws.max_row+1
      ws.cell(row=row_to_write, column=1).value = f"{name}"
      wb.save(filename=filename)
      print(row_to_write)

  def mark_attendance(self,name,date):
    row_num  = None
    wb,ws = self.get_latest_month()
    for row in ws.iter_rows(min_col= 1, max_col =1):
       print(row[0].value ,name)
       if row[0].value == name:
          
          row_num = row[0].row
          break
    date_col = None
    for col in range(2, ws.max_column + 1):
        print(str(ws.cell(row=1, column=col).value).split(' ')[0], str(date))
        if str(ws.cell(row=1, column=col).value).split(' ')[0]== date:
            
            date_col = col
            break
    print(name, date)
    ws.cell(row=row_num, column=date_col).value = "P"
    wb.save(self.filename)
      


